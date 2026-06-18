from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from core.models import PerfilOrganizacional, SetorOrganizacional
from core.services.permissoes_organizacionais import usuario_tem_acesso_global
from rh.models import Funcionario, Vaga, VinculoAvaliacaoDesempenho


COLUNAS_OBRIGATORIAS = [
    'Nome',
    'Admissao',
    'Descricao cargo',
    'Descricao Ccusto',
    'Descricao Servico',
    'Descricao Dpto',
    'Gestor',
    'usuario no sistema',
]

SETOR_MAP = {
    'COMERCIAL': 'CA',
    'COMPRAS': 'CO',
    'DIRETORIA': 'DI',
    'FINANCEIRO': 'FI',
    'OBRA': 'OB',
    'PRODUCAO': 'FA',
    'PROJETOS': 'PR',
    'QUALIDADE': 'QA',
    'RECURSOS HUMANOS': 'RH',
    'TECNOLOGIA DA INFORMACAO': 'TI',
}

IGNORAR_NOMES = {'DE', 'DA', 'DO', 'DAS', 'DOS', 'E'}
CPF_PLACEHOLDER_INICIAL = 90000000001


@dataclass
class LinhaEmpregado:
    numero: int
    nome: str
    admissao: Any
    cargo: str
    centro_custo: str
    servico: str
    departamento: str
    gestor: str
    usuario_planilha: str
    nome_normalizado: str
    setor_codigo: str | None = None
    data_admissao: date | None = None
    usuario: Any = None
    funcionario: Funcionario | None = None
    email_resolvido: str = ''
    login_resolvido: str = ''


class RelatorioImportacao:
    def __init__(self) -> None:
        self.contadores = Counter()
        self.erros: list[str] = []
        self.setores_nao_reconhecidos: list[str] = []
        self.gestores_nao_encontrados: list[str] = []
        self.conflitos_setor: list[str] = []
        self.emails_gerados: list[str] = []
        self.vinculos_admin: list[str] = []
        self.overrides: list[str] = []
        self.possiveis_duplicidades: list[str] = []
        self.datas_futuras: list[str] = []
        self.cargos_vazios: list[str] = []

    def inc(self, chave: str, valor: int = 1) -> None:
        self.contadores[chave] += valor

    def add(self, lista: str, mensagem: str) -> None:
        getattr(self, lista).append(mensagem)


def valor_texto(valor: Any) -> str:
    if valor is None:
        return ''
    texto = str(valor).strip()
    if texto.lower() in {'nan', 'none', 'nat'}:
        return ''
    return re.sub(r'\s+', ' ', texto)


def remover_acentos(texto: str) -> str:
    return unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')


def normalizar_texto(texto: Any) -> str:
    texto = remover_acentos(valor_texto(texto))
    texto = re.sub(r'\s+', ' ', texto)
    return texto.upper().strip()


def normalizar_login(valor: Any) -> str:
    login = valor_texto(valor).lower()
    login = re.sub(r'\s+', '', login)
    if login.startswith('@'):
        login = login[1:]
    return login


def gerar_email_por_nome(nome: str, dominio: str) -> str:
    nome_limpo = remover_acentos(nome)
    partes = [
        re.sub(r'[^a-z0-9]', '', parte.lower())
        for parte in re.split(r'\s+', nome_limpo.strip())
    ]
    partes = [parte for parte in partes if parte and parte.upper() not in IGNORAR_NOMES]
    if not partes:
        local = 'usuario'
    elif len(partes) == 1:
        local = partes[0]
    else:
        local = f'{partes[0]}.{partes[-1]}'
    return f'{local}@{dominio}'


def dividir_nome(nome: str) -> tuple[str, str]:
    partes = [parte for parte in valor_texto(nome).title().split(' ') if parte]
    if not partes:
        return '', ''
    if len(partes) == 1:
        return partes[0], ''
    return partes[0], ' '.join(partes[1:])


def datas_ciclo(valor: Any, relatorio: RelatorioImportacao, linha: int, nome: str) -> date:
    data_lida: date | None = None
    if isinstance(valor, datetime):
        data_lida = valor.date()
    elif isinstance(valor, date):
        data_lida = valor
    else:
        texto = valor_texto(valor)
        for formato in ('%d/%m/%Y', '%Y-%m-%d'):
            try:
                data_lida = datetime.strptime(texto, formato).date()
                break
            except ValueError:
                continue

    if not data_lida:
        relatorio.add('erros', f'Linha {linha}: data de admissao invalida para {nome}; usando data atual.')
        return timezone.localdate()

    if data_lida > timezone.localdate():
        relatorio.inc('datas_futuras')
        relatorio.add('datas_futuras', f'Linha {linha}: {nome} com admissao futura em {data_lida.isoformat()}.')

    return data_lida


class ImportadorEmpregadosAvaliacao:
    def __init__(
        self,
        *,
        caminho: Path,
        senha_padrao: str,
        dominio_email: str,
        atualizar_senha_existentes: bool,
        forcar_troca_senha_existentes: bool,
        permitir_vinculo_admin: bool,
        relatorio: RelatorioImportacao,
    ) -> None:
        self.caminho = caminho
        self.senha_padrao = senha_padrao
        self.dominio_email = dominio_email.strip().lower()
        self.atualizar_senha_existentes = atualizar_senha_existentes
        self.forcar_troca_senha_existentes = forcar_troca_senha_existentes
        self.permitir_vinculo_admin = permitir_vinculo_admin
        self.relatorio = relatorio
        self.User = get_user_model()
        self.linhas: list[LinhaEmpregado] = []
        self.linhas_por_nome: dict[str, LinhaEmpregado] = {}
        self.linhas_por_primeiro_nome: dict[str, LinhaEmpregado] = {}
        self.linhas_por_login: dict[str, LinhaEmpregado] = {}
        self.cpfs_reservados = set(Funcionario.objects.values_list('cpf', flat=True))

    def executar(self) -> RelatorioImportacao:
        self.linhas = self._ler_planilha()
        self._preparar_indices_planilha()
        for linha in self.linhas:
            self._processar_funcionario(linha)
        for linha in self.linhas:
            self._processar_vinculo(linha)
        return self.relatorio

    def _ler_planilha(self) -> list[LinhaEmpregado]:
        if not self.caminho.exists():
            raise CommandError(f'Arquivo nao encontrado: {self.caminho}')

        wb = load_workbook(self.caminho, data_only=True)
        ws = wb.active
        cabecalho_original = [valor_texto(cell.value) for cell in ws[1]]
        mapa_colunas = {normalizar_texto(coluna): idx for idx, coluna in enumerate(cabecalho_original)}

        faltantes = [
            coluna
            for coluna in COLUNAS_OBRIGATORIAS
            if normalizar_texto(coluna) not in mapa_colunas
        ]
        if faltantes:
            raise CommandError(f'Colunas obrigatorias ausentes: {", ".join(faltantes)}. Encontradas: {cabecalho_original}')

        linhas: list[LinhaEmpregado] = []
        for numero, valores in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(valor_texto(valor) for valor in valores):
                self.relatorio.inc('linhas_ignoradas')
                continue

            self.relatorio.inc('linhas_lidas')
            nome = valor_texto(valores[mapa_colunas[normalizar_texto('Nome')]])
            if not nome:
                self.relatorio.inc('linhas_ignoradas')
                self.relatorio.add('erros', f'Linha {numero}: nome ausente.')
                continue

            cargo = valor_texto(valores[mapa_colunas[normalizar_texto('Descricao cargo')]])
            if not cargo:
                cargo = 'Nao informado'
                self.relatorio.inc('cargos_vazios')
                self.relatorio.add('cargos_vazios', f'Linha {numero}: {nome}.')

            departamento = valor_texto(valores[mapa_colunas[normalizar_texto('Descricao Dpto')]])
            setor_codigo = SETOR_MAP.get(normalizar_texto(departamento))
            if not setor_codigo:
                setor_codigo = Vaga.SETORES.COMERCIAL
                self.relatorio.add('setores_nao_reconhecidos', f'Linha {numero}: {nome} - "{departamento}" usando CA.')

            linha = LinhaEmpregado(
                numero=numero,
                nome=nome,
                admissao=valores[mapa_colunas[normalizar_texto('Admissao')]],
                cargo=cargo,
                centro_custo=valor_texto(valores[mapa_colunas[normalizar_texto('Descricao Ccusto')]]),
                servico=valor_texto(valores[mapa_colunas[normalizar_texto('Descricao Servico')]]),
                departamento=departamento,
                gestor=valor_texto(valores[mapa_colunas[normalizar_texto('Gestor')]]),
                usuario_planilha=valor_texto(valores[mapa_colunas[normalizar_texto('usuario no sistema')]]),
                nome_normalizado=normalizar_texto(nome),
                setor_codigo=setor_codigo,
                data_admissao=datas_ciclo(
                    valores[mapa_colunas[normalizar_texto('Admissao')]],
                    self.relatorio,
                    numero,
                    nome,
                ),
            )
            linhas.append(linha)
            self.relatorio.inc('linhas_validas')

        duplicados = [nome for nome, total in Counter(linha.nome_normalizado for linha in linhas).items() if total > 1]
        for nome in duplicados:
            self.relatorio.inc('possiveis_duplicidades')
            self.relatorio.add('possiveis_duplicidades', f'Nome duplicado na planilha: {nome}.')

        return linhas

    def _preparar_indices_planilha(self) -> None:
        self.linhas_por_nome = {linha.nome_normalizado: linha for linha in self.linhas}

        por_primeiro_nome: defaultdict[str, list[LinhaEmpregado]] = defaultdict(list)
        for linha in self.linhas:
            partes = linha.nome_normalizado.split()
            if partes:
                por_primeiro_nome[partes[0]].append(linha)
            login = normalizar_login(linha.usuario_planilha)
            if login:
                self.linhas_por_login[login] = linha
                if '@' not in login:
                    self.linhas_por_login[f'{login}@{self.dominio_email}'] = linha

        self.linhas_por_primeiro_nome = {
            primeiro_nome: linhas[0]
            for primeiro_nome, linhas in por_primeiro_nome.items()
            if len(linhas) == 1
        }

    def _processar_funcionario(self, linha: LinhaEmpregado) -> None:
        usuario, criado, senha_definida = self._buscar_ou_criar_usuario(linha)
        linha.usuario = usuario
        if criado:
            self.relatorio.inc('usuarios_criados')
        else:
            self.relatorio.inc('usuarios_existentes_vinculados')
        if senha_definida:
            self.relatorio.inc('usuarios_com_senha_padrao_definida')

        funcionario, criado_funcionario = self._criar_ou_atualizar_funcionario(linha)
        linha.funcionario = funcionario
        if criado_funcionario:
            self.relatorio.inc('funcionarios_criados')
        else:
            self.relatorio.inc('funcionarios_atualizados')

        self._sincronizar_perfil_organizacional(linha)
        self.relatorio.inc('setores_associados')

    def _buscar_ou_criar_usuario(self, linha: LinhaEmpregado) -> tuple[Any, bool, bool]:
        login_planilha = normalizar_login(linha.usuario_planilha)
        email_gerado = gerar_email_por_nome(linha.nome, self.dominio_email)
        login_para_criar = login_planilha
        email_para_criar = email_gerado

        if login_planilha:
            usuario = self._buscar_usuario_por_login(login_planilha)
            if login_planilha == 'admin':
                if usuario:
                    self.relatorio.add(
                        'vinculos_admin',
                        f'Linha {linha.numero}: {linha.nome} tentou vincular ao usuario admin.',
                    )
                    if not self.permitir_vinculo_admin:
                        login_para_criar = ''
                        usuario = None
                else:
                    self.relatorio.add(
                        'vinculos_admin',
                        f'Linha {linha.numero}: {linha.nome} tentou criar/vincular login admin.',
                    )
                    if not self.permitir_vinculo_admin:
                        login_para_criar = ''

            if usuario:
                linha.email_resolvido = usuario.email
                linha.login_resolvido = usuario.username
                senha_redefinida = self._aplicar_politica_usuario_existente(usuario)
                return usuario, False, senha_redefinida

            if login_para_criar:
                if '@' in login_para_criar:
                    email_para_criar = login_para_criar
                    username_base = login_para_criar.split('@', 1)[0]
                else:
                    email_para_criar = f'{login_para_criar}@{self.dominio_email}'
                    username_base = login_para_criar
            else:
                usuario = self._buscar_usuario_por_email_ou_username(email_gerado)
                if usuario:
                    linha.email_resolvido = usuario.email
                    linha.login_resolvido = usuario.username
                    senha_redefinida = self._aplicar_politica_usuario_existente(usuario)
                    return usuario, False, senha_redefinida
                username_base = email_gerado.split('@', 1)[0]
        else:
            usuario = self._buscar_usuario_por_email_ou_username(email_gerado)
            if usuario:
                linha.email_resolvido = usuario.email
                linha.login_resolvido = usuario.username
                senha_redefinida = self._aplicar_politica_usuario_existente(usuario)
                return usuario, False, senha_redefinida
            username_base = email_gerado.split('@', 1)[0]

        username = self._username_disponivel(username_base)
        if username != username_base:
            self.relatorio.inc('possiveis_duplicidades')
            self.relatorio.add(
                'possiveis_duplicidades',
                f'Linha {linha.numero}: username {username_base} ja existia; usando {username}.',
            )

        first_name, last_name = dividir_nome(linha.nome)
        usuario = self.User(
            username=username,
            email=email_para_criar,
            first_name=first_name,
            last_name=last_name,
            is_staff=False,
            is_superuser=False,
        )
        usuario.set_password(self.senha_padrao)
        usuario.must_change_password = True
        usuario.save()
        self.relatorio.inc('usuarios_novos_marcados_troca_senha')

        linha.email_resolvido = usuario.email
        linha.login_resolvido = usuario.username
        self.relatorio.add('emails_gerados', f'Linha {linha.numero}: {linha.nome} -> {usuario.username} / {usuario.email}.')
        return usuario, True, True

    def _buscar_usuario_por_login(self, login: str) -> Any | None:
        buscas = Q(username__iexact=login)
        if '@' in login:
            buscas |= Q(email__iexact=login)
        else:
            buscas |= Q(email__iexact=f'{login}@{self.dominio_email}')
        return self.User.objects.filter(buscas).first()

    def _buscar_usuario_por_email_ou_username(self, email: str) -> Any | None:
        username = email.split('@', 1)[0]
        return self.User.objects.filter(Q(email__iexact=email) | Q(username__iexact=username)).first()

    def _aplicar_politica_usuario_existente(self, usuario: Any) -> bool:
        update_fields = []
        senha_redefinida = False

        if self.atualizar_senha_existentes:
            usuario.set_password(self.senha_padrao)
            update_fields.append('password')
            self.relatorio.inc('senhas_existentes_redefinidas')
            senha_redefinida = True

        if self.atualizar_senha_existentes or self.forcar_troca_senha_existentes:
            usuario.must_change_password = True
            update_fields.append('must_change_password')
            self.relatorio.inc('usuarios_existentes_marcados_troca_senha')

        if update_fields:
            usuario.save(update_fields=update_fields)

        return senha_redefinida

    def _username_disponivel(self, base: str) -> str:
        base = re.sub(r'[^a-z0-9._-]', '', remover_acentos(base.lower())).strip('._-') or 'usuario'
        username = base[:150]
        if not self.User.objects.filter(username__iexact=username).exists():
            return username
        for contador in range(2, 1000):
            sufixo = f'.{contador}'
            candidato = f'{base[:150 - len(sufixo)]}{sufixo}'
            if not self.User.objects.filter(username__iexact=candidato).exists():
                return candidato
        raise IntegrityError(f'Nao foi possivel gerar username unico para {base}.')

    def _criar_ou_atualizar_funcionario(self, linha: LinhaEmpregado) -> tuple[Funcionario, bool]:
        funcionario = Funcionario.objects.filter(usuario=linha.usuario).first()
        if not funcionario:
            funcionario = Funcionario.objects.filter(nome_completo__iexact=linha.nome).first()

        criado = funcionario is None
        if criado:
            funcionario = Funcionario(cpf=self._proximo_cpf_placeholder())

        funcionario.usuario = linha.usuario
        funcionario.nome_completo = linha.nome
        funcionario.data_admissao = linha.data_admissao or timezone.localdate()
        funcionario.cargo = linha.cargo or 'Nao informado'
        funcionario.salario = funcionario.salario or Decimal('0.00')
        funcionario.situacao = Funcionario.SITUACAO.ATIVO
        funcionario.setor = linha.setor_codigo or Vaga.SETORES.COMERCIAL
        funcionario.departamento_planilha = linha.departamento
        funcionario.centro_custo_planilha = linha.centro_custo
        funcionario.servico_planilha = linha.servico
        funcionario.gestor_nome_planilha = linha.gestor
        funcionario.usuario_planilha = linha.usuario_planilha
        funcionario.save()
        return funcionario, criado

    def _proximo_cpf_placeholder(self) -> str:
        proximo = CPF_PLACEHOLDER_INICIAL
        while str(proximo) in self.cpfs_reservados:
            proximo += 1
        cpf = str(proximo)
        self.cpfs_reservados.add(cpf)
        return cpf

    def _sincronizar_perfil_organizacional(self, linha: LinhaEmpregado) -> None:
        setor = SetorOrganizacional.objects.get(codigo=linha.setor_codigo)
        PerfilOrganizacional.objects.update_or_create(
            usuario=linha.usuario,
            defaults={
                'setor': setor,
                'cargo': linha.cargo,
                'data_admissao': linha.data_admissao,
                'pode_ser_avaliado': True,
                'ativo': True,
            },
        )

    def _processar_vinculo(self, linha: LinhaEmpregado) -> None:
        gestor_valor = self._gestor_com_override(linha)
        gestor_usuario = None
        gestor_funcionario = None
        gestor_email = ''
        setor_gestor = ''

        if gestor_valor:
            linha_gestor = self._resolver_linha_gestor(gestor_valor)
            if linha_gestor and linha_gestor.usuario:
                gestor_usuario = linha_gestor.usuario
                gestor_funcionario = linha_gestor.funcionario
                gestor_email = gestor_usuario.email
                setor_gestor = linha_gestor.setor_codigo or ''
                self.relatorio.inc('gestores_encontrados_por_usuario')
                if gestor_funcionario:
                    self.relatorio.inc('gestores_encontrados_por_funcionario')
            else:
                self.relatorio.inc('gestores_nao_encontrados')
                self.relatorio.add('gestores_nao_encontrados', f'Linha {linha.numero}: {linha.nome} -> gestor "{gestor_valor}".')
        else:
            self.relatorio.inc('gestores_nao_encontrados')
            self.relatorio.add('gestores_nao_encontrados', f'Linha {linha.numero}: {linha.nome} sem gestor informado.')

        if gestor_usuario and gestor_usuario == linha.usuario:
            self.relatorio.add('erros', f'Linha {linha.numero}: gestor igual ao avaliado para {linha.nome}; vinculo sem gestor.')
            gestor_usuario = None
            gestor_funcionario = None
            gestor_email = ''
            setor_gestor = ''

        if gestor_usuario and setor_gestor and setor_gestor != linha.setor_codigo and not usuario_tem_acesso_global(gestor_usuario):
            self.relatorio.inc('conflitos_setor')
            self.relatorio.add(
                'conflitos_setor',
                f'Linha {linha.numero}: {linha.nome} setor {linha.setor_codigo} com gestor {gestor_usuario.username} setor {setor_gestor}.',
            )

        vinculo = VinculoAvaliacaoDesempenho.objects.filter(avaliado=linha.funcionario, ativo=True).first()
        criado = vinculo is None
        if criado:
            vinculo = VinculoAvaliacaoDesempenho(avaliado=linha.funcionario, ativo=True)

        vinculo.gestor_usuario = gestor_usuario
        vinculo.gestor_funcionario = gestor_funcionario
        vinculo.gestor_nome_planilha = gestor_valor
        vinculo.gestor_email_resolvido = gestor_email
        vinculo.setor_avaliado = linha.setor_codigo
        vinculo.setor_gestor = setor_gestor
        vinculo.origem = 'PLANILHA'
        vinculo.save()

        self._atualizar_gestor_direto(linha, gestor_usuario)
        self.relatorio.inc('vinculos_avaliacao_criados' if criado else 'vinculos_avaliacao_atualizados')

    def _gestor_com_override(self, linha: LinhaEmpregado) -> str:
        return valor_texto(linha.gestor)

    def _resolver_linha_gestor(self, gestor_valor: str) -> LinhaEmpregado | None:
        gestor_norm = normalizar_texto(gestor_valor)
        if not gestor_norm:
            return None

        if gestor_norm in self.linhas_por_nome:
            return self.linhas_por_nome[gestor_norm]

        login = normalizar_login(gestor_valor)
        if login in self.linhas_por_login:
            return self.linhas_por_login[login]

        partes = gestor_norm.split()
        if len(partes) == 1 and partes[0] in self.linhas_por_primeiro_nome:
            return self.linhas_por_primeiro_nome[partes[0]]

        candidatos = [linha for linha in self.linhas if linha.nome_normalizado.startswith(gestor_norm)]
        if len(candidatos) == 1:
            return candidatos[0]

        candidatos = [linha for linha in self.linhas if gestor_norm in linha.nome_normalizado]
        if len(candidatos) == 1:
            return candidatos[0]

        if len(candidatos) > 1:
            self.relatorio.inc('possiveis_duplicidades')
            self.relatorio.add('possiveis_duplicidades', f'Gestor ambiguo "{gestor_valor}": {[c.nome for c in candidatos]}.')
        return None

    def _atualizar_gestor_direto(self, linha: LinhaEmpregado, gestor_usuario: Any | None) -> None:
        PerfilOrganizacional.objects.filter(usuario=linha.usuario).update(gestor_direto=gestor_usuario)


class Command(BaseCommand):
    help = 'Importa empregados ativos para a base de avaliacao de desempenho.'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', type=str)
        parser.add_argument('--dry-run', action='store_true', help='Simula a importacao sem gravar no banco.')
        parser.add_argument('--senha-padrao', default='tmg@2026')
        parser.add_argument('--atualizar-senha-existentes', action='store_true')
        parser.add_argument('--forcar-troca-senha-existentes', action='store_true')
        parser.add_argument('--dominio-email', default='i9tmg.com.br')
        parser.add_argument('--permitir-vinculo-admin', action='store_true')

    def handle(self, *args, **options):
        relatorio = RelatorioImportacao()
        caminho = Path(options['arquivo']).expanduser()

        importador = ImportadorEmpregadosAvaliacao(
            caminho=caminho,
            senha_padrao=options['senha_padrao'],
            dominio_email=options['dominio_email'],
            atualizar_senha_existentes=options['atualizar_senha_existentes'],
            forcar_troca_senha_existentes=options['forcar_troca_senha_existentes'],
            permitir_vinculo_admin=options['permitir_vinculo_admin'],
            relatorio=relatorio,
        )

        with transaction.atomic():
            relatorio = importador.executar()
            if options['dry_run']:
                transaction.set_rollback(True)

        self._imprimir_relatorio(relatorio, dry_run=options['dry_run'])

    def _imprimir_relatorio(self, relatorio: RelatorioImportacao, *, dry_run: bool) -> None:
        titulo = 'DRY-RUN concluido; nenhuma alteracao foi gravada.' if dry_run else 'Importacao concluida.'
        self.stdout.write(self.style.SUCCESS(titulo))

        ordem_contadores = [
            ('linhas_lidas', 'Total de linhas lidas'),
            ('linhas_validas', 'Total de linhas validas'),
            ('linhas_ignoradas', 'Total de linhas ignoradas'),
            ('funcionarios_criados', 'Total de funcionarios criados'),
            ('funcionarios_atualizados', 'Total de funcionarios atualizados'),
            ('usuarios_criados', 'Total de usuarios criados'),
            ('usuarios_existentes_vinculados', 'Total de usuarios existentes vinculados'),
            ('usuarios_com_senha_padrao_definida', 'Total de usuarios com senha padrao definida'),
            ('usuarios_novos_marcados_troca_senha', 'Total de usuarios novos marcados para troca de senha'),
            ('usuarios_existentes_marcados_troca_senha', 'Total de usuarios existentes marcados para troca de senha'),
            ('senhas_existentes_redefinidas', 'Total de senhas de existentes redefinidas'),
            ('setores_associados', 'Total de setores associados'),
            ('vinculos_avaliacao_criados', 'Total de vinculos de avaliacao criados'),
            ('vinculos_avaliacao_atualizados', 'Total de vinculos de avaliacao atualizados'),
            ('gestores_encontrados_por_usuario', 'Total de gestores encontrados por usuario'),
            ('gestores_encontrados_por_funcionario', 'Total de gestores encontrados por funcionario'),
            ('gestores_nao_encontrados', 'Total de gestores nao encontrados'),
            ('conflitos_setor', 'Total de conflitos de setor entre gestor e avaliado'),
            ('possiveis_duplicidades', 'Total de possiveis duplicidades'),
            ('datas_futuras', 'Total de datas futuras'),
            ('cargos_vazios', 'Total de cargos vazios'),
        ]
        for chave, rotulo in ordem_contadores:
            self.stdout.write(f'{rotulo}: {relatorio.contadores[chave]}')

        listas = [
            ('erros', 'Erros por linha'),
            ('setores_nao_reconhecidos', 'Setores nao reconhecidos'),
            ('gestores_nao_encontrados', 'Gestores nao encontrados'),
            ('conflitos_setor', 'Conflitos gestor/setor'),
            ('emails_gerados', 'E-mails/logins gerados'),
            ('vinculos_admin', 'Possiveis vinculos com usuario admin'),
            ('overrides', 'Overrides aplicados'),
            ('possiveis_duplicidades', 'Possiveis duplicidades'),
            ('datas_futuras', 'Datas futuras'),
            ('cargos_vazios', 'Cargos vazios'),
        ]
        for atributo, titulo in listas:
            itens = getattr(relatorio, atributo)
            if not itens:
                continue
            self.stdout.write(self.style.WARNING(f'\n{titulo}:'))
            for item in itens:
                self.stdout.write(f'- {item}')
